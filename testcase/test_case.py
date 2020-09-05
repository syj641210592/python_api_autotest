from com_func.test_log import log
from com_func.confread import config
from com_func.sql_request import mysql
from com_func.com_request_func import com_request, com_assertEqual, com_excel_read
from jsonpath import jsonpath
import unittest
import ddt
import openpyxl

@ddt.ddt
class TestWithdraw(unittest.TestCase):
    '''提现接口测试'''
    Worksheet_name = "withdraw"
    excel, data_list = com_excel_read(Worksheet_name)
    leave_amount_max = config.get("PRESET", "leave_amount_max")  # 获取提取上限

    @classmethod
    def setUpClass(cls):
        '''登录获取token,用户id'''
        cls.excel.wb = openpyxl.load_workbook(cls.excel.path)   # 重新载入workbook
        params = config.get("PRESET", "loan_params")
        respone = com_request(TestWithdraw, "login", params=params)
        cls.token = "Bearer" + " " + jsonpath(respone, "$..token")[0]
        cls.member_id = jsonpath(respone, "$..id")[0]

    def setUp(self):
        pass

    @ddt.data(*data_list)
    @ddt.unpack
    def test_case(self, info, **kwargs):
        '''{info}'''
        try:
            if kwargs["sql_cheack"]:  # 存在sql校验,获取初期余额
                res = mysql.sql_read(kwargs["sql_cheack"], self.member_id)
                amount_prime = res["leave_amount"]
            respone = com_request(TestWithdraw, self.Worksheet_name, **kwargs, token=self.token)
            if kwargs["sql_cheack"]:  # 存在sql校验,获取提现后余额
                res = mysql.sql_read(kwargs["sql_cheack"], self.member_id)
                amount_new = res["leave_amount"]
                amount = amount_prime - amount_new
                print(f"实际{info[:2]}:", float(amount))
                print(f"预期{info[:2]}:", float(self.params["amount"]))
                self.assertEqual(float(amount), float(self.params["amount"]))  # 校验提现的余额实际与预期
            com_assertEqual(self, respone, eval(kwargs["expect"]))
        except AssertionError as e:
            print(e)
            log.error(f"用例--{info}--执行失败", exc_info=True)
            self.excel.excel_write(self.Worksheet_name, kwargs["id"], "失败")
            raise e
        else:
            log.info(f"用例--{info}--执行成功", exc_info=False)
            self.excel.excel_write(self.Worksheet_name, kwargs["id"], "成功")

    def tearDown(self):
        pass

    @classmethod
    def tearDownClass(cls):
        cls.excel.wb.save(cls.excel.path)
        cls.excel.wb.close()


@ddt.ddt
class TestRecharge(unittest.TestCase):
    '''充值接口测试'''
    Worksheet_name = "recharge"
    excel, data_list = com_excel_read(Worksheet_name)

    @classmethod
    def setUpClass(cls):
        '''登录获取token,用户id'''
        cls.excel.wb = openpyxl.load_workbook(cls.excel.path)   # 重新载入workbook
        params = config.get("PRESET", "loan_params")
        respone = com_request(TestRecharge, "login", params=params)
        cls.token = "Bearer" + " " + jsonpath(respone, "$..token")[0]
        cls.member_id = jsonpath(respone, "$..id")[0]

    def setUp(self):
        pass

    @ddt.data(*data_list)
    @ddt.unpack
    def test_case(self, info, **kwargs):
        '''{info}'''
        try:
            if kwargs["sql_cheack"]:  # 存在sql校验,获取初期余额
                res = mysql.sql_read(kwargs["sql_cheack"], self.member_id)
                amount_prime = res["leave_amount"]
            respone = com_request(TestRecharge, self.Worksheet_name, **kwargs, token=self.token)
            if kwargs["sql_cheack"]:  # 存在sql校验,获取充值后余额
                res = mysql.sql_read(kwargs["sql_cheack"], self.member_id)
                amount_new = res["leave_amount"]
                amount = amount_new - amount_prime
                print("实际充值:", float(amount))
                print("预期充值:", float(self.params["amount"]))
                self.assertEqual(float(amount), float(self.params["amount"]))  # 校验充值的余额实际与预期
            com_assertEqual(self, respone, eval(kwargs["expect"]))
        except AssertionError as e:
            log.error(f"用例--{info}--执行失败", exc_info=True)
            self.excel.excel_write(self.Worksheet_name, kwargs["id"], "失败")
            raise e
        else:
            log.info(f"用例--{info}--执行成功", exc_info=False)
            self.excel.excel_write(self.Worksheet_name, kwargs["id"], "成功")

    def tearDown(self):
        pass

    @classmethod
    def tearDownClass(cls):
        cls.excel.wb.save(cls.excel.path)
        cls.excel.wb.close()


@ddt.ddt
class TestRegister(unittest.TestCase):
    '''注册接口测试'''
    Worksheet_name = "register"
    excel, data_list = com_excel_read(Worksheet_name)

    @classmethod
    def setUpClass(cls):
        cls.excel.wb = openpyxl.load_workbook(cls.excel.path) 

    def setUp(self):
        pass
    
    @ddt.data(*data_list)
    @ddt.unpack
    def test_case(self, info, **kwargs):
        '''{info}'''
        try:
            respone = com_request(TestRegister, self.Worksheet_name, **kwargs)
            com_assertEqual(self, respone, eval(kwargs["expect"]))
        except AssertionError as e:
            log.error(f"用例--{info}--执行失败", exc_info=True)
            self.excel.excel_write(self.Worksheet_name, kwargs["id"], "失败")
            raise e
        else:
            log.info(f"用例--{info}--执行成功", exc_info=False)
            self.excel.excel_write(self.Worksheet_name, kwargs["id"], "成功")

    def tearDown(self):
        pass

    @classmethod
    def tearDownClass(cls):
        cls.excel.wb.save(cls.excel.path)
        cls.excel.wb.close()


@ddt.ddt
class TestLogin(unittest.TestCase):
    '''登录接口测试'''
    Worksheet_name = "login"
    excel, data_list = com_excel_read(Worksheet_name)

    @classmethod
    def setUpClass(cls):
        cls.excel.wb = openpyxl.load_workbook(cls.excel.path) 

    def setUp(self):
        pass
    
    @ddt.data(*data_list)
    @ddt.unpack
    def test_case(self, info, **kwargs):
        '''{info}'''
        try:
            respone = com_request(TestLogin, self.Worksheet_name, **kwargs)
            com_assertEqual(self, respone, eval(kwargs["expect"]))
        except AssertionError as e:
            log.error(f"用例--{info}--执行失败", exc_info=True)
            self.excel.excel_write(self.Worksheet_name, kwargs["id"], "失败")
            raise e
        else:
            log.info(f"用例--{info}--执行成功", exc_info=False)
            self.excel.excel_write(self.Worksheet_name, kwargs["id"], "成功")

    def tearDown(self):
        pass

    @classmethod
    def tearDownClass(cls):
        cls.excel.wb.save(cls.excel.path)
        cls.excel.wb.close()
